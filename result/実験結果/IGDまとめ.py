#!/usr/bin/env python
# -*- coding: utf-8 -*-
import numpy
import glob
import sys;
import openpyxl as px
from openpyxl.styles import Alignment
import os;
from openpyxl.styles import Border, Side
import string
import numpy as np
import glob 
import re

def numericalSort(value):
    numbers = re.compile(r'(\d+)')
    parts = numbers.split(value)
    parts[1::2] = map(int, parts[1::2])
    return parts


def getAllFileName(str_,target):
	return sorted(glob.iglob(str_+'/**/' + target, recursive=True), key=numericalSort);



if __name__ == "__main__":
	filename = "AVE_STD.dat";
	FILEDEMILITER = "\t";

	FILES = getAllFileName("result\\NSGA2",filename);
	ALGORITHMSET = [];
	PROBLEMSET = [];
	OBJECTIVESET = [];
	SFSET = [];


	#　store all position
	OBJECTHASH = {};
	ALGORITHMHASH = {};
	PROBLEMHASH = {};

	Alpharbet = list(string.ascii_uppercase);

################################dataの確認 
	for file in FILES:
		if (not file.split("\\")[1] in ALGORITHMSET):
			ALGORITHMSET.append(file.split("\\")[1]);

		if (not file.split("\\")[2] in PROBLEMSET):
			PROBLEMSET.append(file.split("\\")[2]);
			
		if (not file.split("\\")[3] in OBJECTIVESET):
			OBJECTIVESET.append(file.split("\\")[3]);
		
		if (not file.split("\\")[5] in SFSET):
			print(file.split("\\")[5]);
			SFSET.append(file.split("\\")[5]);

	
	print("ALGO"+"	"+str(len(ALGORITHMSET)))
	print("PROBLEM"+"	"+str(len(PROBLEMSET)))
	print("OBJECTIV"+"	"+str(len(OBJECTIVESET)))
	print("SFSET" +"	" +str(len(SFSET)));
	input();	

#	if(os.path.exists("result.xlsx")):
#		print("上書きする可能性があるため終了します");
#		input();

	wb = px.Workbook();
	
	ws = wb.get_sheet_by_name("Sheet");

	d = 0;
	for obj in OBJECTIVESET:
		OBJECTHASH[str(obj)] =  d;
		d = d+1;


	counter = 0;
	for algo in ALGORITHMSET:
		for sf in SFSET:
			counter=counter + 1;
	#		print(Alpharbet[2*counter]+str(1)+":"+Alpharbet[2*counter+1]+str(1));
			ws.merge_cells(Alpharbet[2*counter]+str(1)+":"+Alpharbet[2*counter+1]+str(1));
			ws[Alpharbet[2*counter]+str(1)].alignment = Alignment( wrap_text=True, horizontal='center', vertical='center')
			ws.cell(row = 1,column = 2*counter+1).value = algo +"	"+sf;
			ALGORITHMHASH[str(algo) +"	"+str(sf)] =  Alpharbet[2*counter];
			ALGORITHMHASH[str(algo) +"	"+str(sf)+"ave"] =  Alpharbet[2*counter+1];

			ws.cell(row = 2,column = 2*counter+1).value = "ave";
			ws.cell(row = 2,column = 2*counter+2).value = "std";

	for algo in ALGORITHMSET:
		counter = 3;
		for problem in PROBLEMSET:
			
			upper = counter;
			PROBLEMHASH[problem] = upper;
			for obj in OBJECTIVESET:
				ws['A'+str(counter)].value = problem;
				ws['B'+str(counter)].value = (str(str(obj).split("OBJ")[0]));
				counter = counter + 1;
			ws.merge_cells('A'+str(upper)+':A'+str(counter-1));
			ws['A'+str(upper)].alignment = Alignment( wrap_text=True, horizontal='center', vertical='center')
			

	d = getAllFileName("result",filename);
	for file in d:
		filetext = file.split("\\");
		algorithmkey = filetext[1] + "	" + filetext[5];
		PROBELMKEY = filetext[2]
		OBJECTKEY = filetext[3];
#		print(PROBLEMHASH[str(PROBELMKEY)]);
#		print(OBJECTHASH[str(OBJECTKEY)]);
		first = -10000;
		second = -10000;

		with open(file,"r") as fin:
			data = fin.readlines();
			data = data[len(data)-1];
			data = data.replace("\n","");
			first = data.split(FILEDEMILITER)[0]
			print(first)
			second = data.split(FILEDEMILITER)[1]
			
		POSITION = ALGORITHMHASH[algorithmkey]+str(PROBLEMHASH[str(PROBELMKEY)] +OBJECTHASH[str(OBJECTKEY)] );
		ws[POSITION].value = float(first);
		POSITION = ALGORITHMHASH[algorithmkey+"ave"]+str(PROBLEMHASH[str(PROBELMKEY)] +OBJECTHASH[str(OBJECTKEY)] );
		ws[POSITION].value = float(second);

#

#		OBJECTIVE = file.split("\\")
		




	#table = np.loadtxt('numbers.dat')

	wb.save("result.xlsx");


	
